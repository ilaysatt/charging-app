"""Generate a standalone Excel charging tracker template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

wb = openpyxl.Workbook()

# --- Styles ---
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
title_font = Font(bold=True, color="1E3A5F", size=14)
subtitle_font = Font(bold=True, color="1E3A5F", size=11)
input_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
calc_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
setting_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def style_range(ws, min_row, max_row, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border


# =====================
# SETTINGS sheet
# =====================
ws_settings = wb.active
ws_settings.title = "Settings"
ws_settings.sheet_properties.tabColor = "FFC107"

ws_settings["A1"] = "Charging Tracker - Settings"
ws_settings["A1"].font = title_font
ws_settings.merge_cells("A1:C1")

ws_settings["A3"] = "Setting"
ws_settings["B3"] = "Value"
ws_settings["C3"] = "Description"
style_header_row(ws_settings, 3, 3)

settings_data = [
    ("kWh per 1%", 2.05, "How many kWh per 1% of battery"),
    ("Price per kWh", 0.35, "Cost per kilowatt-hour"),
]
for i, (name, val, desc) in enumerate(settings_data, start=4):
    ws_settings.cell(row=i, column=1, value=name).border = thin_border
    cell = ws_settings.cell(row=i, column=2, value=val)
    cell.fill = setting_fill
    cell.border = thin_border
    cell.number_format = "0.00"
    ws_settings.cell(row=i, column=3, value=desc).border = thin_border

ws_settings.column_dimensions["A"].width = 18
ws_settings.column_dimensions["B"].width = 12
ws_settings.column_dimensions["C"].width = 35

# Named ranges for easy reference
dn1 = DefinedName("KWH_PER_PCT", attr_text="Settings!$B$4")
dn2 = DefinedName("PRICE_PER_KWH", attr_text="Settings!$B$5")
wb.defined_names.add(dn1)
wb.defined_names.add(dn2)

# =====================
# CHARGES sheet
# =====================
ws = wb.create_sheet("Charges")
ws.sheet_properties.tabColor = "2196F3"
wb.active = wb.sheetnames.index("Charges")

ws["A1"] = "Charging Log"
ws["A1"].font = title_font
ws.merge_cells("A1:H1")

# Column headers
headers = ["Date", "VIN", "Start %", "End %", "kWh per %", "kWh Charged", "Cost", "Notes"]
for col, h in enumerate(headers, 1):
    ws.cell(row=3, column=col, value=h)
style_header_row(ws, 3, len(headers))

# Pre-fill 200 rows with formulas
MAX_ROWS = 200
for row in range(4, 4 + MAX_ROWS):
    # Date column - user input
    ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
    ws.cell(row=row, column=1).fill = input_fill
    # VIN - user input
    ws.cell(row=row, column=2).fill = input_fill
    # Start % - user input
    ws.cell(row=row, column=3).fill = input_fill
    ws.cell(row=row, column=3).number_format = "0.0"
    # End % - user input
    ws.cell(row=row, column=4).fill = input_fill
    ws.cell(row=row, column=4).number_format = "0.0"
    # kWh per % - defaults to setting, user can override per row
    ws.cell(row=row, column=5).value = "=Settings!$B$4"
    ws.cell(row=row, column=5).fill = input_fill
    ws.cell(row=row, column=5).number_format = "0.00"
    # kWh Charged = (End% - Start%) * kWh per %
    ws.cell(row=row, column=6).value = (
        '=IF(AND(C%d<>"",D%d<>""),(D%d-C%d)*E%d,"")' % (row, row, row, row, row)
    )
    ws.cell(row=row, column=6).fill = calc_fill
    ws.cell(row=row, column=6).number_format = "0.00"
    # Cost = kWh * price
    ws.cell(row=row, column=7).value = (
        '=IF(F%d<>"",F%d*Settings!$B$5,"")' % (row, row)
    )
    ws.cell(row=row, column=7).fill = calc_fill
    ws.cell(row=row, column=7).number_format = "0.00"
    # Notes - user input
    ws.cell(row=row, column=8).fill = input_fill

    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

# Column widths
col_widths = [14, 20, 10, 10, 12, 14, 12, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Legend
legend_row = 4 + MAX_ROWS + 1
ws.cell(row=legend_row, column=1, value="Legend:").font = subtitle_font
ws.cell(row=legend_row + 1, column=1, value="Blue cells").fill = input_fill
ws.cell(row=legend_row + 1, column=2, value="= You fill in")
ws.cell(row=legend_row + 2, column=1, value="Green cells").fill = calc_fill
ws.cell(row=legend_row + 2, column=2, value="= Auto-calculated (don't edit)")

# =====================
# SUMMARY sheet
# =====================
ws_sum = wb.create_sheet("Summary")
ws_sum.sheet_properties.tabColor = "4CAF50"

ws_sum["A1"] = "Charging Summary"
ws_sum["A1"].font = title_font
ws_sum.merge_cells("A1:E1")

# --- Per VIN summary ---
ws_sum["A3"] = "Summary by VIN"
ws_sum["A3"].font = subtitle_font

sum_headers = ["VIN", "Sessions", "Total kWh", "Total Cost"]
for col, h in enumerate(sum_headers, 1):
    ws_sum.cell(row=4, column=col, value=h)
style_header_row(ws_sum, 4, len(sum_headers))

# Use UNIQUE + SUMPRODUCT to summarize by VIN (works in Excel 365/2021+)
# For broader compatibility, we'll use a simpler approach with placeholder VIN rows
# that users fill in
ws_sum["A5"] = "(Enter VIN below to see totals)"
ws_sum["A5"].font = Font(italic=True, color="888888")

for row in range(6, 16):
    ws_sum.cell(row=row, column=1).fill = input_fill
    ws_sum.cell(row=row, column=1).border = thin_border
    # Sessions count
    ws_sum.cell(row=row, column=2).value = (
        '=IF(A%d<>"",COUNTIF(Charges!B$4:B$203,A%d),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=2).fill = calc_fill
    ws_sum.cell(row=row, column=2).border = thin_border
    # Total kWh
    ws_sum.cell(row=row, column=3).value = (
        '=IF(A%d<>"",SUMIF(Charges!B$4:B$203,A%d,Charges!F$4:F$203),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=3).fill = calc_fill
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = "0.00"
    # Total Cost
    ws_sum.cell(row=row, column=4).value = (
        '=IF(A%d<>"",SUMIF(Charges!B$4:B$203,A%d,Charges!G$4:G$203),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=4).fill = calc_fill
    ws_sum.cell(row=row, column=4).border = thin_border
    ws_sum.cell(row=row, column=4).number_format = "0.00"

# Grand totals
ws_sum.cell(row=17, column=1, value="GRAND TOTAL").font = Font(bold=True)
ws_sum.cell(row=17, column=1).border = thin_border
ws_sum.cell(row=17, column=2).value = '=COUNTA(Charges!A4:A203)'
ws_sum.cell(row=17, column=2).fill = calc_fill
ws_sum.cell(row=17, column=2).border = thin_border
ws_sum.cell(row=17, column=3).value = "=SUM(Charges!F4:F203)"
ws_sum.cell(row=17, column=3).fill = calc_fill
ws_sum.cell(row=17, column=3).border = thin_border
ws_sum.cell(row=17, column=3).number_format = "0.00"
ws_sum.cell(row=17, column=4).value = "=SUM(Charges!G4:G203)"
ws_sum.cell(row=17, column=4).fill = calc_fill
ws_sum.cell(row=17, column=4).border = thin_border
ws_sum.cell(row=17, column=4).number_format = "0.00"

# --- Monthly summary ---
ws_sum["A19"] = "Summary by Month"
ws_sum["A19"].font = subtitle_font

month_headers = ["Month (YYYY-MM)", "Sessions", "Total kWh", "Total Cost"]
for col, h in enumerate(month_headers, 1):
    ws_sum.cell(row=20, column=col, value=h)
style_header_row(ws_sum, 20, len(month_headers))

ws_sum["A21"] = "(Enter month below, e.g. 2026-03)"
ws_sum["A21"].font = Font(italic=True, color="888888")

for row in range(22, 35):
    ws_sum.cell(row=row, column=1).fill = input_fill
    ws_sum.cell(row=row, column=1).border = thin_border
    # Sessions - count rows where TEXT(date,"YYYY-MM") matches
    ws_sum.cell(row=row, column=2).value = (
        '=IF(A%d<>"",SUMPRODUCT((TEXT(Charges!A$4:A$203,"YYYY-MM")=A%d)*1),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=2).fill = calc_fill
    ws_sum.cell(row=row, column=2).border = thin_border
    # Total kWh
    ws_sum.cell(row=row, column=3).value = (
        '=IF(A%d<>"",SUMPRODUCT((TEXT(Charges!A$4:A$203,"YYYY-MM")=A%d)*Charges!F$4:F$203),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=3).fill = calc_fill
    ws_sum.cell(row=row, column=3).border = thin_border
    ws_sum.cell(row=row, column=3).number_format = "0.00"
    # Total Cost
    ws_sum.cell(row=row, column=4).value = (
        '=IF(A%d<>"",SUMPRODUCT((TEXT(Charges!A$4:A$203,"YYYY-MM")=A%d)*Charges!G$4:G$203),"")' % (row, row)
    )
    ws_sum.cell(row=row, column=4).fill = calc_fill
    ws_sum.cell(row=row, column=4).border = thin_border
    ws_sum.cell(row=row, column=4).number_format = "0.00"

# Column widths
ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 14

# Freeze panes on Charges sheet
ws.freeze_panes = "A4"

# Save
output_path = "charging_tracker.xlsx"
wb.save(output_path)
print("Created: %s" % output_path)
